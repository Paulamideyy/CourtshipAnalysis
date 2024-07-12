import os
import csv
import openpyxl

# Paths to folders and workbook
first_csv_folder = r"E:\Paulami\Ground-truthing Data\Double Blind\Shrinivas\Attempted Copulation\manual_output_files"
second_csv_folder = r"E:\Paulami\Ground-truthing Data\Double Blind\Shrinivas\Attempted Copulation\JAABA_output"
workbook_path = r"E:\Paulami\Ground-truthing Data\Double Blind\Shrinivas\Attempted Copulation\AC_DoubleBlind.xlsx"

# Open the workbook
workbook = openpyxl.load_workbook(workbook_path)

# Print out all sheet names in the workbook for debugging
print("Available sheet names in the workbook:")
print(workbook.sheetnames)

# Function to write data to the workbook
def write_data_to_sheet(sheet, data, start_row, start_col, skip_header=False):
    for row_index, row in enumerate(data):
        if skip_header and row_index == 0:
            continue  # Skip the header row
        for col_index, value in enumerate(row):
            cell = sheet.cell(row=start_row + row_index, column=start_col + col_index)
            try:
                cell.value = float(value)
                cell.number_format = 'General'
            except ValueError:
                cell.value = value
            except Exception as e:
                print(f"Error writing value '{value}' to cell {cell.coordinate}: {e}")

# Get inputs for JAABA filenames to strip 
striping_JAABA_name1 = input("Enter the JAABA score file name to be removed: ")
striping_JAABA_name2 = input("Enter the JAABA score file name to be removed: ")

# Iterate through CSV files in the first folder
for csv_file in os.listdir(first_csv_folder):
    if csv_file.endswith(".csv"):
        csv_path = os.path.join(first_csv_folder, csv_file)
        csv_filename, _ = os.path.splitext(csv_file)
        sheet_name = csv_filename.replace("_output", "")  # Strip "_output" from CSV file name
        
        # Print derived sheet name for debugging
        print(f"Derived sheet name from first folder: {sheet_name}")
        
        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read data from CSV
            with open(csv_path, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)

            # Write data to the sheet starting from the second row
            write_data_to_sheet(sheet, data, start_row=2, start_col=1, skip_header=False)

        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")

# Iterate through CSV files in the second folder
for csv_file in os.listdir(second_csv_folder):
    if csv_file.endswith(".csv"):
        csv_path = os.path.join(second_csv_folder, csv_file)
        csv_filename, _ = os.path.splitext(csv_file)
        sheet_name = csv_filename.replace(striping_JAABA_name1, "").replace(striping_JAABA_name2, "")  # Strip any JAABA score file name from the final output name. 
        
        # sheet_name = csv_filename.replace("_output", "")  # Strip "_output" from CSV file name
        
        # Print derived sheet name for debugging
        print(f"Derived sheet name from second folder: {sheet_name}")
        
        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read data from CSV
            with open(csv_path, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)

            # Write data to the 5th and 6th columns in the sheet, starting from the second row
            write_data_to_sheet(sheet, data, start_row=2, start_col=3, skip_header=False)

        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")

# Save the workbook
workbook.save(workbook_path)
